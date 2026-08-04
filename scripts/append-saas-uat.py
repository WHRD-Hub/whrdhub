#!/usr/bin/env python3
"""
Append a "SAAS Test Scenarios" tab to the EXISTING WHRD Hub UAT workbook,
preserving all sheets already in the file (including any manual edits made after
the original generator ran). Idempotent: re-running replaces only the SAAS tab.

Usage:  python3 scripts/append-saas-uat.py
"""

import os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(HERE, "..", "public", "WHRD-Hub-UAT-Worksheet.xlsx")
SHEET = "SAAS Test Scenarios"

# palette (matches the generator)
PURPLE, PURPLE_DARK = "5B21B6", "3B0764"
GREY_DARK, WHITE = "1F2937", "FFFFFF"
GREEN_L, GREEN = "DCFCE7", "15803D"
RED_L, RED = "FEE2E2", "991B1B"
BLUE_L, BLUE = "DBEAFE", "1D4ED8"
GREY_L, GREY_M = "F9FAFB", "6B7280"
INPUT_BG = "FFFEF7"
BORDER = "E5E7EB"

thin = Side(style="thin", color=BORDER)
box = Border(top=thin, bottom=thin, left=thin, right=thin)

def font(bold=False, sz=10, color=GREY_DARK, italic=False):
    return Font(name="Calibri", bold=bold, size=sz, color=color, italic=italic)
def fill(rgb):
    return PatternFill("solid", fgColor=rgb)
def al(h="left", v="top", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

TYPE_COLORS = {"Positive": (GREEN_L, GREEN), "Negative": (RED_L, RED), "Validation": (BLUE_L, BLUE)}

COLS = ["Test ID","Module","Scenario Title","Test Type","Pre-Conditions","Test Data Required",
        "Test Steps (numbered)","Expected Result","Actual Result\n(Tester fills in)",
        "Status\n(Pass/Fail/Blocked)","Defect Ref","Tester Name","Date Tested","Comments /\nObservations"]
WIDTHS = [10,16,32,11,34,28,52,42,30,14,12,14,13,28]

# module -> list of (id, title, type, pre, data, steps, expected)
MODULES = [
 ("SAAS-01 — Access, Shared Auth & Cross-Platform", [
  ("SS-001","Landing page loads (split view)","Positive","SAAS staging URL. Desktop Chrome.","No credentials.",
   "1. Open the SAAS URL.\n2. Let the page fully load.\n3. Observe layout on desktop.",
   "Landing loads within 5s: official content on the left, the live Community Feed on the right, the full WHRD Hub logo (with 'Protect. Heal. Nurture.'), and the tagline 'A home for Women Human Rights Defenders across Kenya & Beyond'. No console errors."),
  ("SS-002","Sign in with shared account","Positive","An account created on the reporting platform exists (shared auth DB).","Registered email + password.",
   "1. Go to the SAAS /login.\n2. Enter the email/password used on the reporting platform.\n3. Click Sign In.",
   "The same credentials work (shared Supabase auth). User signs in and lands in the SAAS dashboard, not the marketing landing."),
  ("SS-003","Continue with Google lands inside the SAAS","Positive","Google provider enabled; Supabase redirect URLs include the SAAS callback (wildcard).","A Google account.",
   "1. On /login click 'Continue with Google'.\n2. Pick the Google account.\n3. Observe the destination.",
   "After Google auth the user returns to the SAAS and lands in the dashboard (or /onboarding for a new account) — not the landing page."),
  ("SS-004","Switch to reporting platform","Positive","Signed in on the SAAS.","None.",
   "1. In the top nav, click 'Reporting'.\n2. Observe navigation.",
   "User is taken to the reporting platform dashboard. The session carries over (same auth)."),
  ("SS-005","Admin view toggle persists","Positive","Account has is_hub_admin = true.","Admin account.",
   "1. Sign in as the admin.\n2. Open the profile/user menu; use the Admin view toggle to enter the Hub console.\n3. Sign out and sign back in.",
   "The admin can switch between member and admin views. The chosen view persists across sign-out/in on the same device."),
 ]),
 ("SAAS-02 — Onboarding (T&C, County, CBO, Femtorship)", [
  ("SS-006","New user routed to onboarding, T&C first","Validation","Brand-new account; onboarding not completed.","New account.",
   "1. Sign in / sign up with a new account.\n2. Observe the first screen.",
   "User is sent to onboarding before the dashboard. Terms & Conditions are presented first."),
  ("SS-007","Cannot proceed without accepting T&C","Validation","On onboarding.","None.",
   "1. Leave the T&C acceptance unchecked.\n2. Try to continue.",
   "The continue button stays disabled until the T&C checkbox is ticked."),
  ("SS-008","County + organization (CBO) capture","Positive","On onboarding.","County: Nairobi. CBO: pick existing or 'Add new'.",
   "1. Select a county network.\n2. Choose an existing CBO or create a new one.\n3. Continue.",
   "County and CBO are saved to the profile/membership. Creating a new CBO adds it as pending verification."),
  ("SS-009","Femtorship questionnaire captured","Positive","On onboarding / profile.","Sample answers to the femtorship questions.",
   "1. Complete the femtorship questions (leadership, guidance areas, can provide support, etc.).\n2. Save.",
   "Answers are stored on the femtorship profile and used later for matching."),
  ("SS-010","Onboarding not re-triggered","Validation","Onboarding completed once.","Same account.",
   "1. Sign out.\n2. Sign back in.",
   "User goes straight to the dashboard; onboarding is not shown again."),
 ]),
 ("SAAS-03 — Landing & Marketing Pages", [
  ("SS-011","Tagline shows '& Beyond'","Positive","Landing loaded.","None.",
   "1. Read the hero tagline and footer.",
   "Reads 'A home for Women Human Rights Defenders across Kenya & Beyond' on the landing and footer."),
  ("SS-012","County Networks page lists 9","Positive","On the public site.","None.",
   "1. Open Our Work → County Networks.\n2. Count the networks; open one.",
   "Nine county networks are listed; each opens its own page with content (no 404)."),
  ("SS-013","Our Impact section loads","Positive","On the public site.","None.",
   "1. About Us → Our Impact.",
   "The Impact section opens with key figures and highlights (no dead link, not a repeat of another page)."),
  ("SS-014","Blog list + article + gallery","Positive","At least one published story.","None.",
   "1. Open Voices → Blog.\n2. Open a story.",
   "Blog list renders; the article shows the full body, and any photos load as a gallery after the text."),
  ("SS-015","Footer, socials and logo","Positive","Any public page.","None.",
   "1. Scroll to the footer.",
   "Footer shows the full logo, contact details, and working social links. No 'Report Abuse' CTA in the footer."),
 ]),
 ("SAAS-04 — Community Feed", [
  ("SS-016","Feed shows posts + Hub videos","Positive","Feed has approved posts and Hub videos.","None.",
   "1. Open the Community Feed.\n2. Scroll.",
   "Approved posts and Hub YouTube videos appear; pinned items are at the top, then newest first."),
  ("SS-017","Guest reaction limit + sign-in gate","Validation","Signed out.","None.",
   "1. As a guest, react/support on posts.\n2. Continue past 3 reactions.",
   "Guests can react up to 3 times; the 4th prompts sign-in. The 3 guest reactions sync to the account after signing in."),
  ("SS-018","Signed-in reaction updates count","Positive","Signed in.","None.",
   "1. Support a post.\n2. Observe the count.",
   "The support/reaction count increments immediately and persists on refresh."),
  ("SS-019","YouTube link post embeds player","Positive","A post created with a YouTube link.","None.",
   "1. Open a post that has a YouTube link.",
   "The video renders as an embedded, playable player (not a raw link)."),
  ("SS-020","Reporting entry present","Positive","Any page.","None.",
   "1. Look for the route back to the reporting platform.",
   "A clear path to the reporting platform is available in the nav/menu."),
 ]),
 ("SAAS-05 — Story Authoring (Editor & Workflow)", [
  ("SS-021","Create post with media","Positive","Signed in.","An image or document under the size limit.",
   "1. Open the composer.\n2. Write an update and attach media.\n3. Post.",
   "The post is created; a member's post goes to review, an admin's publishes immediately."),
  ("SS-022","Write a story with the rich editor","Positive","Signed in.","Sample story text.",
   "1. Composer → Write a story.\n2. Use the toolbar: heading, list, quote, link, image.\n3. Preview.",
   "The TipTap editor applies formatting; Preview matches how the story will render."),
  ("SS-023","Save draft","Positive","Writing a story.","Draft content.",
   "1. Click 'Save draft'.\n2. Go to Profile → My Activity → Stories.",
   "The story is saved as a Draft and appears under activity with an Edit link."),
  ("SS-024","Submit for review","Positive","A draft or new story.","Story with enough content.",
   "1. Click 'Submit for review'.",
   "Status becomes 'In review' and the Hub admins are notified. It appears in the admin queue."),
  ("SS-025","Revise a declined story","Positive","A story that was declined with a note.","Revised content.",
   "1. Profile → Stories → open the declined story ('Revise').\n2. Read the decline note, edit, resubmit.",
   "The decline note is shown; edits save and resubmitting returns it to 'In review'."),
 ]),
 ("SAAS-06 — Hub Admin Console", [
  ("SS-026","Overview + needs-attention inbox","Positive","Signed in as Hub admin; items pending.","Admin account.",
   "1. Open the Hub overview.\n2. Toggle the inbox between Posts / CBOs / Stories.",
   "KPIs and management cards render; the inbox toggles between pending Posts, CBOs and Stories, each row opening its detail."),
  ("SS-027","Review, edit & publish","Positive","A pending post/story exists.","None.",
   "1. Open a pending item.\n2. Edit it in the editor.\n3. Approve & publish.",
   "Edits save; on publish the item goes live on the feed/blog and the author is notified."),
  ("SS-028","Decline with note","Positive","A pending item.","Decline reason.",
   "1. Open the item.\n2. Decline with a note.",
   "Item is marked declined, the author is notified with the note, and it shows under the Declined filter."),
  ("SS-029","Verify a CBO","Positive","A CBO is pending verification.","None.",
   "1. CBOs → open an unverified CBO.\n2. Use the verify action at the top.",
   "The CBO becomes Verified; its members and content can then go public."),
  ("SS-030","Member drill-down + femtorship","Positive","A member with content/matches.","None.",
   "1. Members → open a member.\n2. Review their content table and femtorship panel.",
   "The member's posts/stories are listed and their femtorship role and matches (mentoring / mentored by) are shown."),
  ("SS-031","Admin composer: pin + YouTube","Positive","Signed in as admin.","A YouTube link.",
   "1. Composer → enable 'Pin to top', paste a YouTube link with text.\n2. Post.",
   "The admin post auto-publishes, is pinned to the top of the feed, and shows the embedded video."),
 ]),
 ("SAAS-07 — Profile, Femtorship & Accessibility", [
  ("SS-032","Edit profile","Positive","Signed in.","New name/title/bio/county.",
   "1. Profile → Account.\n2. Update fields and save.",
   "Changes save and are reflected across the platform."),
  ("SS-033","Femtorship form + matches","Positive","Signed in.","Updated answers.",
   "1. Profile → Femtorship.\n2. Update answers; view any connected matches.",
   "Answers save; existing matches are listed, or the form shows if none yet."),
  ("SS-034","Accessibility controls persist","Positive","Signed in or guest.","None.",
   "1. Open Accessibility (profile or widget).\n2. Change text size / contrast / reduce motion.\n3. Reload.",
   "Settings apply immediately and persist on this device after reload."),
  ("SS-035","Privacy: delete own content / account","Validation","Signed in with own content.","None.",
   "1. Profile → Privacy & Security.\n2. Delete a post/story; review the delete-account option.",
   "Own content is removed; the delete-account flow warns clearly before any irreversible action."),
  ("SS-036","Activity tabs","Positive","Signed in with activity.","None.",
   "1. Profile → My Activity.\n2. Switch Posts / Stories / Supported.",
   "Each tab lists the user's items with correct status pills; declined/draft stories offer edit."),
 ]),
]

wb = load_workbook(XLSX)  # preserves existing sheets + their formatting
if SHEET in wb.sheetnames:
    del wb[SHEET]
ws = wb.create_sheet(SHEET)

ncols = len(COLS)
r = 1
# Title
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
c = ws.cell(r, 1, "WHRD HUB — SAAS COMMUNITY PLATFORM — UAT TEST SCENARIOS")
c.font = font(True, 14, WHITE); c.fill = fill(PURPLE_DARK); c.alignment = al("center", "center")
ws.row_dimensions[r].height = 32
r += 1
# Header
for i, h in enumerate(COLS, 1):
    hc = ws.cell(r, i, h)
    hc.font = font(True, 10, WHITE); hc.fill = fill(GREY_DARK); hc.alignment = al("center", "center"); hc.border = box
ws.row_dimensions[r].height = 40
r += 1

def input_cell(cell):
    cell.fill = fill(INPUT_BG)
    cell.border = Border(top=thin, bottom=Side(style="dashed", color="FCD34D"), left=thin, right=thin)
    cell.alignment = al()

for mod_title, scenarios in MODULES:
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
    mc = ws.cell(r, 1, mod_title)
    mc.font = font(True, 11, WHITE); mc.fill = fill(PURPLE); mc.alignment = al("left", "center", False); mc.border = box
    ws.row_dimensions[r].height = 22
    r += 1
    for (tid, title, ttype, pre, data, steps, expected) in scenarios:
        row_vals = [tid, mod_title.split(" — ")[0], title, ttype, pre, data, steps, expected, "", "Not Tested", "", "", "", ""]
        for i, v in enumerate(row_vals, 1):
            cell = ws.cell(r, i, v)
            cell.border = box
            if i == 1:
                cell.font = font(True, 10, PURPLE); cell.alignment = al()
            elif i == 2:
                cell.font = font(True, 10, WHITE); cell.fill = fill(PURPLE); cell.alignment = al("center", "center")
            elif i == 3:
                cell.font = font(True, 10, GREY_DARK); cell.alignment = al()
            elif i == 4:
                bg, fg = TYPE_COLORS.get(ttype, (GREY_L, GREY_DARK))
                cell.font = font(True, 10, fg); cell.fill = fill(bg); cell.alignment = al("center", "center")
            elif i == 10:
                cell.font = font(True, 10, GREY_M); cell.fill = fill(GREY_L); cell.alignment = al("center", "center")
            elif i in (9, 11, 12, 13, 14):
                input_cell(cell)
            else:
                cell.font = font(False, 10, GREY_DARK); cell.alignment = al()
        # rough row height by longest text
        longest = max(len(str(x)) for x in row_vals)
        ws.row_dimensions[r].height = 150 if longest > 400 else 110 if longest > 200 else 70 if longest > 100 else 34
        r += 1

for i, w in enumerate(WIDTHS, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A3"

# place the SAAS tab right after the reporting "Test Scenarios" tab if present
try:
    names = wb.sheetnames
    if "Test Scenarios" in names:
        target = names.index("Test Scenarios") + 1
        wb.move_sheet(SHEET, offset=(target - names.index(SHEET)))
except Exception:
    pass

wb.save(XLSX)
total = sum(len(s) for _, s in MODULES)
print(f"OK: added '{SHEET}' with {total} scenarios across {len(MODULES)} modules.")
print("Sheets now:", wb.sheetnames)
